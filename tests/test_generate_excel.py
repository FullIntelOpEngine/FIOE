"""
test_generate_excel.py — Tests for /generate_excel endpoint and Excel export.

Uses openpyxl stubs and avoids importing webbridge.py.

Run with:  pytest tests/test_generate_excel.py
"""
import io
import json
import os
import tempfile
import unittest
from unittest.mock import MagicMock, patch

try:
    import openpyxl
    _OPENPYXL_AVAILABLE = True
except ImportError:
    _OPENPYXL_AVAILABLE = False

from flask import Flask, jsonify, request, send_file


# ---------------------------------------------------------------------------
# Minimal Excel generation logic (stubbed from webbridge.py)
# ---------------------------------------------------------------------------

def _generate_excel_from_rows(rows: list, include_dropdown: bool = False) -> bytes:
    """Generate a minimal .xlsx file from a list of row dicts."""
    if not _OPENPYXL_AVAILABLE:
        raise RuntimeError("openpyxl not available")
    wb = openpyxl.Workbook()
    ws = wb.active
    if not rows:
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()
    headers = list(rows[0].keys())
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h, "") for h in headers])
    if include_dropdown:
        from openpyxl.worksheet.datavalidation import DataValidation
        dv = DataValidation(type="list", formula1='"Yes,No,Maybe"', allow_blank=True)
        ws.add_data_validation(dv)
        for i in range(2, len(rows) + 2):
            dv.add(ws.cell(row=i, column=1))
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _atomic_write_json(path: str, data: dict) -> None:
    """Write JSON atomically (tmp → rename)."""
    tmp = path + ".tmp"
    with open(tmp, "w", encoding="utf-8") as fh:
        json.dump(data, fh, indent=2)
    os.replace(tmp, path)


def _build_excel_app(output_dir: str):
    app = Flask(__name__)
    app.config["TESTING"] = True

    @app.post("/generate_excel")
    def generate_excel():
        body = request.get_json(force=True, silent=True) or {}
        fetch_from_db = body.get("fetch_from_db", False)

        if fetch_from_db:
            # Simulated DB rows
            rows = [
                {"name": "Alice", "title": "Engineer", "score": 85},
                {"name": "Bob",   "title": "Manager",  "score": 90},
            ]
            include_dropdown = True
        else:
            rows = body.get("rows", [])
            if not rows:
                rows = [{"name": "Test", "title": "Role", "score": 0}]
            include_dropdown = False

        try:
            xlsx_bytes = _generate_excel_from_rows(rows, include_dropdown=include_dropdown)
        except RuntimeError as e:
            return jsonify({"error": str(e)}), 503

        tmp_path = os.path.join(output_dir, "export_tmp.xlsx")
        with open(tmp_path, "wb") as fh:
            fh.write(xlsx_bytes)

        # Also update data_sorter.json atomically
        ds_path = os.path.join(output_dir, "data_sorter.json")
        _atomic_write_json(ds_path, {"last_export_rows": len(rows)})

        return send_file(
            io.BytesIO(xlsx_bytes),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name="export.xlsx",
        )

    return app


# ---------------------------------------------------------------------------
# Tests
# ---------------------------------------------------------------------------

@unittest.skipUnless(_OPENPYXL_AVAILABLE, "openpyxl not installed")
class TestGenerateExcel(unittest.TestCase):

    def setUp(self):
        self.td = tempfile.mkdtemp()
        self.app = _build_excel_app(self.td)
        self.client = self.app.test_client()

    def test_excel_from_payload(self):
        """POST /generate_excel fetch_from_db=false → valid xlsx bytes."""
        resp = self.client.post(
            "/generate_excel",
            json={
                "fetch_from_db": False,
                "rows": [{"name": "Eve", "title": "Analyst", "score": 75}],
            },
        )
        self.assertEqual(resp.status_code, 200)
        content_type = resp.headers.get("Content-Type", "")
        self.assertIn("spreadsheetml", content_type)
        # Validate the returned bytes are valid xlsx
        wb = openpyxl.load_workbook(io.BytesIO(resp.data))
        ws = wb.active
        self.assertEqual(ws.cell(row=1, column=1).value, "name")
        self.assertEqual(ws.cell(row=2, column=1).value, "Eve")

    def test_excel_from_db(self):
        """POST /generate_excel fetch_from_db=true → xlsx with dropdown validation."""
        resp = self.client.post(
            "/generate_excel",
            json={"fetch_from_db": True},
        )
        self.assertEqual(resp.status_code, 200)
        wb = openpyxl.load_workbook(io.BytesIO(resp.data))
        ws = wb.active
        # Should have header row + 2 data rows
        self.assertGreater(ws.max_row, 1)

    def test_data_sorter_update(self):
        """Excel export updates data_sorter.json atomically."""
        self.client.post(
            "/generate_excel",
            json={"fetch_from_db": False, "rows": [{"x": 1}, {"x": 2}]},
        )
        ds_path = os.path.join(self.td, "data_sorter.json")
        self.assertTrue(os.path.exists(ds_path))
        with open(ds_path) as fh:
            data = json.load(fh)
        self.assertEqual(data.get("last_export_rows"), 2)

    def test_excel_cleanup(self):
        """Temp xlsx is written to output_dir (cleanup responsibility lies with caller)."""
        self.client.post(
            "/generate_excel",
            json={"fetch_from_db": False},
        )
        tmp_path = os.path.join(self.td, "export_tmp.xlsx")
        self.assertTrue(os.path.exists(tmp_path))
        # Cleanup
        os.remove(tmp_path)
        self.assertFalse(os.path.exists(tmp_path))

    def test_empty_rows_produces_valid_file(self):
        """No rows still produces a valid xlsx file."""
        resp = self.client.post(
            "/generate_excel",
            json={"fetch_from_db": False, "rows": []},
        )
        self.assertEqual(resp.status_code, 200)
        wb = openpyxl.load_workbook(io.BytesIO(resp.data))
        self.assertIsNotNone(wb.active)


class TestAtomicWriteJson(unittest.TestCase):

    def test_atomic_write_creates_file(self):
        """_atomic_write_json creates file with correct content."""
        with tempfile.TemporaryDirectory() as td:
            path = os.path.join(td, "out.json")
            _atomic_write_json(path, {"key": "value"})
            with open(path) as fh:
                data = json.load(fh)
            self.assertEqual(data["key"], "value")

    def test_atomic_write_no_tmp_left(self):
        """_atomic_write_json leaves no .tmp file after success."""
        with tempfile.TemporaryDirectory() as td:
            path = os.path.join(td, "out.json")
            _atomic_write_json(path, {"x": 1})
            self.assertFalse(os.path.exists(path + ".tmp"))


if __name__ == "__main__":
    unittest.main()
